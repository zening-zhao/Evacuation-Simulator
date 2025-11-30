# LayoutBuilder.py
# this module will read the layout file (could be in excel or csv format), and build out the field
import openpyxl
from openpyxl import Workbook
from openpyxl.styles.colors import COLOR_INDEX
from ThemeColorConverter import ThemeColorConverter
from LayoutCell import LayoutCell
import constants
import math as math
import numpy as np
import CellularAutomata as CellularAutomata
import EvacuationZoneCell as EvacuationZoneCell
import operator
import logging
import random
from logger_config import setup_logger
from PySide6.QtGui import QPainter, QPen, QPixmap, QIcon, QAction, QColor, QBrush
from PySide6.QtWidgets import QApplication, QGraphicsScene, QGraphicsView, QGraphicsLineItem, QGraphicsEllipseItem, QGraphicsRectItem
from PySide6.QtCore import Qt, QLineF, QRectF

setup_logger()
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


#---------------------------------------------------------------------------
##  Component to build out the evacuation layout
#---------------------------------------------------------------------------
class LayoutBuilderQt():

    def __init__(self, path, canvas, config):
        self.path = path
        self.LayoutMap = []                 # will be a list of LayoutCell[]. each item of this list is another list of layoutcell
        self.AutomataList = []              # a list of cellular automata.
        self.min_row = 0
        self.min_column = 0
        self.max_row = 0
        self.max_column = 0
        self.canvas = canvas
        self.max_pedestrian_capicity = 0
        self.config = config                # pass the configuration instance here
        self.evacuation_zone_info = []      # setup a list to store the evacuation zone info
                                            # each item in this list is the EvacuationZoneCell instance

        # predefine a list of directions for further usage
        self.direction_List = [self.config.DIRECTION_CENTER,
                                self.config.DIRECTION_DOWN,
                                self.config.DIRECTION_DOWNLEFT,
                                self.config.DIRECTION_DOWNRIGHT,
                                self.config.DIRECTION_LEFT,
                                self.config.DIRECTION_RIGHT,
                                self.config.DIRECTION_UP,
                                self.config.DIRECTION_UPLEFT,
                                self.config.DIRECTION_UPRIGHT]
        self.Exit_Dict = {}     # store all exit information. will be id and layoutcell list pair
                                # {1:[layoutcell1, layoutcell2, layoutcell3], 2:[layoutcell4, layoutcell5, layoutcell6]}

    #---------------------------------------------------------------------------
    ##  load the input layout file and construct the 2 dimension layout matrix
    #---------------------------------------------------------------------------
    def load_layout_file(self, FORMAT='xlsx'):
        # TODO: will support more files

        layout_workbook = openpyxl.load_workbook(self.path)
        theme_color = ThemeColorConverter(layout_workbook)

        # Loop through each sheet in the workbook
        for sheet in layout_workbook.sheetnames:
            ws = layout_workbook[sheet]
        # Loop through each row and column in the sheet
        # find out the dimension of current defined layout
        # max row, max column, min row (not in white color), min column (not in white color)

        for row in ws.iter_rows():
            currentLayoutRow = []
            for cell in row:
            #newcell = new_ws.cell(row=cell.row, column=cell.column)
                try:
                    color_str = self.get_cell_color(cell, theme_color)
                except:
                    # TODO need more specific error handling. might pop up error message
                    logger.info('Unexpected color')

                if color_str != self.config.LAYOUT_FILE_EMPTY_CELL: # first none empty cell of current row
                    if self.min_row > cell.row or self.min_row == 0:
                        self.min_row = cell.row
                    if self.min_column > cell.column or self.min_column == 0:
                        self.min_column = cell.column
                    if self.max_row < cell.row or self.max_row == 0:
                        self.max_row = cell.row
                    if self.max_column < cell.column or self.max_column == 0:
                        self.max_column = cell.column

                ### TODO will define the color in configuration panel

                if color_str == self.config.LAYOUT_FILE_BORDER:
                    currentCell = LayoutCell(cell.row, cell.column,
                                            self.config.LAYOUT_NUMBER_BORDER,
                                            self.config.LAYOUT_DISPLAY_BORDER,
                                            self.config.LAYOUT_CELL_TYPE_BORDER)
                elif color_str == self.config.LAYOUT_FILE_OBSTACLE:
                    currentCell = LayoutCell(cell.row, cell.column,
                                            self.config.LAYOUT_NUMBER_OBSTACLE,
                                            self.config.LAYOUT_DISPLAY_OBSTACLE,
                                            self.config.LAYOUT_CELL_TYPE_OBSTACLE)
                elif color_str == self.config.LAYOUT_FILE_EXIT:
                    currentCell = LayoutCell(cell.row, cell.column,
                                            self.config.LAYOUT_NUMBER_EXIT,
                                            self.config.LAYOUT_DISPLAY_EXIT,
                                            self.config.LAYOUT_CELL_TYPE_EXIT)
                elif color_str == self.config.LAYOUT_FILE_EMPTY_SPACE:
                    currentCell = LayoutCell(cell.row, cell.column,
                                            self.config.LAYOUT_NUMBER_EMPTY_SPACE,
                                            self.config.LAYOUT_DISPLAY_EMPTY_SPACE,
                                            self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE)
                else: ## Default to empty_CELL
                    currentCell = LayoutCell(cell.row, cell.column,
                                            self.config.LAYOUT_NUMBER_EMPTY_CELL,
                                            self.config.LAYOUT_DISPLAY_EMPTY_CELL,
                                            self.config.LAYOUT_CELL_TYPE_EMPTY_CELL)
                # ## TODO based on number of exit predefined in this layout
                # ## assign a dict to each cell to store the potentials of each
                # ## cell with respect of each exit. (default to -2 as initial value)
                # ## here the exit start with 1.
                # currentCell.static_potential = {key : -2 for key in range(1, self.config.NUMBER_OF_EXIT + 1)}

                currentLayoutRow.append(currentCell)
            self.LayoutMap.append(currentLayoutRow)
            ## at this point we have the full cell information stored in LayoutMap[][]

        logger.debug(f"min_row : {self.min_row} - max_row : {self.max_row} - min_column : {self.min_column} - max_column : {self.max_column}")

    #---------------------------------------------------------------------------
    ##  based on the 2 dimension layout matrix, draw the content to screen
    #---------------------------------------------------------------------------
    def construct_layoutMap(self, canvas):
        ### read the min-max range of the layout map list and draw them on the panel.
        ### LayoutMap is a list of coordiation dict.
        ### LayoutMap[0] -> LayoutRow{}
        ###                 (4,5)-> BLACK
        ###                 (4,6)-> RED
        ###                 (4,7)-> BLUE
        ###                 (4,8)-> WHITE
        ### based on the size of each coordination(x,y)
        ### will draw a square with the adjusted coordination based on predefined CELL size.
        ### The range of the layout map area is defined by the min/max_row, min/mas_column
        # Based on current size calculate how many zone will be assigned vertically.
        vertical_eva_zone_number = math.ceil((self.max_column - self.min_column - 1)
                                             / ( self.config.EVACUATION_ZONE_DIMENSION / self.config.CELL_SIZE))
        horizontal_eva_zone_number = math.ceil((self.max_row - self.min_row - 1)
                                             / ( self.config.EVACUATION_ZONE_DIMENSION / self.config.CELL_SIZE))
        number_of_total_empty_space = 0

        # Draw the map & initialize the layoutMap & automat
        self.AutomataList = []

        for row in self.LayoutMap[self.min_row - 1:]:
            for cell in row[self.min_column - 1:]:
                if cell.type == self.config.LAYOUT_CELL_TYPE_EMPTY_CELL: # skip the empty cell (usually the empty cell is outside of map.)
                    continue
                # for border/obstacle/exit, draw solid rectangle
                if cell.type == self.config.LAYOUT_CELL_TYPE_BORDER or \
                    cell.type == self.config.LAYOUT_CELL_TYPE_OBSTACLE or \
                    cell.type == self.config.LAYOUT_CELL_TYPE_EXIT:

                    pen = QPen(QColor("BLACK"), 1, Qt.DotLine)  # Blue color, 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(cell.color)), Qt.SolidPattern)  # Red solid fill
                    rect = QGraphicsRectItem(QRectF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    self.config.LAYOUT_CELL_SIZE,
                                    self.config.LAYOUT_CELL_SIZE))  # Rectangle with (x, y, width, height)
                    rect.setPen(pen)
                    rect.setBrush(brush)
                    canvas.addItem(rect)
                else:
                # for empty cell/empty space, draw grids only (vertical lines/horizontal line)
                    # Assign zone info to each empty cell.
                    number_of_total_empty_space += 1
                    if cell.type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                        current_assigned_evacuation_zone = int((cell.column - self.min_column - 1) // (self.config.EVACUATION_ZONE_DIMENSION /self.config.CELL_SIZE) + \
                                                            ((cell.row - self.min_row - 1) // (self.config.EVACUATION_ZONE_DIMENSION / self.config.CELL_SIZE)) * vertical_eva_zone_number)
                        self.LayoutMap[cell.row - 1][cell.column - 1].assigned_evacuation_zone = current_assigned_evacuation_zone
                        # Put all empty cell into the big universal celluar automata list (one dimension only)
                        # but for each item in the list, also store the relative coordination info
                        # Here we initialize the Automata with absoulte row/column values.
                        # those values are the universal coordination that can be used to indentify the location of that cell in the map
                        # we might need to tweak a little bit to get the index of them to access the layoutMap array.
                        current_CA = CellularAutomata.CelluarAutomata(cell.row,
                                                                      cell.column,
                                                                      current_assigned_evacuation_zone,
                                                                      current_assigned_evacuation_zone % vertical_eva_zone_number,
                                                                      current_assigned_evacuation_zone // vertical_eva_zone_number)
                        current_CA.occupied = False
                        self.AutomataList.append(current_CA)
                        ### TODO: might get rid of those zone info from each cell but assign them into evacuation_zone_info only?
                        current_ZoneCell = EvacuationZoneCell.EvacuationZoneCell(current_assigned_evacuation_zone // vertical_eva_zone_number,
                                                                                   current_assigned_evacuation_zone % vertical_eva_zone_number,
                                                                                   current_assigned_evacuation_zone,
                                                                                   vertical_eva_zone_number)

                        # Assign initial cell speed to avoid later divid by zero error
                        current_ZoneCell.average_speed = self.config.NORMAL_PEDESTRIAN_SPEED

                        if self.evacuation_zone_info is None:
                            self.evacuation_zone_info = []
                            self.evacuation_zone_info.append(current_ZoneCell)
                        else:
                            if current_ZoneCell not in self.evacuation_zone_info:
                                self.evacuation_zone_info.append(current_ZoneCell)
                    verticle_line = QGraphicsLineItem(QLineF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    cell.column * self.config.LAYOUT_CELL_SIZE + self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE))
                    horizontal_line = QGraphicsLineItem(QLineF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE + self.config.LAYOUT_CELL_SIZE))
                    # Create a pen and brush
                    pen = QPen(QColor("GREY"), 1,  Qt.DotLine)  # Black pen with 1px width, solid line
                    verticle_line.setPen(pen)
                    horizontal_line.setPen(pen)
                    canvas.addItem(verticle_line)
                    canvas.addItem(horizontal_line)

        # at this moment, we have created evacuation zone for current layout. but in extreme case, one evacuation zone could be all
        # obstacle or all empty cell (outside of the layout area). so in this case, that evacuation zone will not even be created.
        # will create dummy evacuation zone there so the entire evacuationzone matrix will be complete matrix.
        # EvaculationZoneCell : assigned_zone_id = {missing IDs}, preferred_exit = -1, number_of_pedestrian = 0, avarage_speed = 1.2

        # find out missing zone ids..
        lst = []
        for zone in self.evacuation_zone_info:
            lst.append(zone.assigned_zone_id)
        missing_numbers = set(range(vertical_eva_zone_number * horizontal_eva_zone_number)) - set(lst)

        for miss_zone_id in missing_numbers:
            current_ZoneCell = EvacuationZoneCell.EvacuationZoneCell(miss_zone_id // vertical_eva_zone_number,
                                                                        miss_zone_id % vertical_eva_zone_number,
                                                                        miss_zone_id,
                                                                        vertical_eva_zone_number)
            current_ZoneCell.obstacle_zone_flag = True # all missing zone are purely obstacle zone.
            self.evacuation_zone_info.append(current_ZoneCell)

        # first Sort the evacuation zone list per assigned_zone_id to reduce potential loop later.
        # in this case, the zone ID can be used to identify the position of the list directly
        self.evacuation_zone_info = sorted(self.evacuation_zone_info, key = lambda x:x.assigned_zone_id)

        # based on current layout, find out all exit and assign exit for each
        # evacuation zone.
        self.initializeExitAndEvacuationZone()


        # debug only: based on the number of empty spaces and cell size and number of people per m^2 calculate the max capicity
        self.max_pedestrian_capicity = number_of_total_empty_space * self.config.CELL_SIZE * self.config.PEDESTRIAN_PER_SQUARE_METER
        logger.info('Max possible number of pedestrian in current layout : '+ str(self.max_pedestrian_capicity))
        logger.info('Current Full size of Automata : ' + str(len(self.AutomataList)))

        # generate random location for the pedestrian
        self.initPedestrian(number=self.config.NUMBER_OF_PEDESTRIAN)

        # plot the pedstrain
        for i in range(len(self.AutomataList)):
            # here adjust the start size and cell size to make the graph more precise
            # We draw the pedestrian inside of each cell. Not cover the border.
            if self.AutomataList[i].occupied: # only setup the spot that is occupied. (not border/exit/obstacle)
                # Create the ellipse item
                ellipse = QGraphicsEllipseItem(QRectF((self.AutomataList[i].y) * self.config.LAYOUT_CELL_SIZE + 1,
                                    (self.AutomataList[i].x) * self.config.LAYOUT_CELL_SIZE + 1,
                                    self.config.LAYOUT_CELL_SIZE - 2,
                                    self.config.LAYOUT_CELL_SIZE - 2))
                if self.AutomataList[i].velocity_mode == 2: # senior people
                    pen = QPen(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_SENIOR_COLOR)), 1, Qt.SolidLine)  # Black pen with 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_SENIOR_COLOR)), Qt.SolidPattern)  # Red solid brush
                else:
                    pen = QPen(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_NON_SENIOR_COLOR)), 1, Qt.SolidLine)  # Black pen with 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_NON_SENIOR_COLOR)), Qt.SolidPattern)  # Red solid brush
                ellipse.setPen(pen)
                ellipse.setBrush(brush)
                canvas.addItem(ellipse)

    #---------------------------------------------------------------------------
    ##  random number generator. will generate amount number of random numbers within [1,max]
    #---------------------------------------------------------------------------
    def randomNumberGenerator(self, amount, max):
        arr = np.empty(amount, dtype=int)
        i = 0
        while True:
            a = int(np.random.uniform(0, max)) # including low but excluding high
            if a not in arr[:i]:
                arr[i] = a
                i += 1
            if i == amount:
                break
        return arr

    #---------------------------------------------------------------------------
    ##  based on the random number, find out which position of cell is occupied
    ##  self.AutomataList has all empty space that can have pedestrian
    ##  the index of this list is randomly chose.
    #---------------------------------------------------------------------------
    def initPedestrian(self, number):

        if number >= self.max_pedestrian_capicity:
            print('Too many pedstrian defined. total number should be under ' + str(self.max_pedestrian_capicity))
        else:
            # Random numbers from a uniform distribution
            arr = self.randomNumberGenerator(number, self.max_pedestrian_capicity)
            # debug:
            logger.debug(arr.shape)

        for i in range(number):
            rnd_value = arr[i] # get random value as index to setup each automata
            try:
                self.AutomataList[rnd_value].occupied = True
                # store the initial step into move_history attribute
                # Here store the absolute location of row/column but not the layout map index
                self.AutomataList[rnd_value].move_history.append(tuple([self.AutomataList[rnd_value].x, self.AutomataList[rnd_value].y]))
                # store the index into corresponding layout cell
                # here use row-1, column -1 to identify the coresponding cell in layout map
                self.LayoutMap[self.AutomataList[rnd_value].x - 1][self.AutomataList[rnd_value].y - 1].automataIndex = rnd_value
                self.AutomataList[rnd_value].current_exit = self.LayoutMap[self.AutomataList[rnd_value].x - 1][self.AutomataList[rnd_value].y - 1].preferred_exit
                ## Apply the SENIOR_PEDESTRIAN_PERCENTAGE
                if np.random.rand() <= self.config.SENIOR_PEDESTRIAN_PERCENTAGE:
                    self.AutomataList[rnd_value].velocity_mode = 2 # half speed senior mode
                    self.AutomataList[rnd_value].last_speed = self.config.SENIOR_PEDESTRIAN_SPEED
                else:
                    self.AutomataList[rnd_value].velocity_mode = 1 # normal speed mode
                    self.AutomataList[rnd_value].last_speed = self.config.NORMAL_PEDESTRIAN_SPEED
            except:
                pass
                # interesting case. ideally the full empty list should cover all rnd value as index.
                # TODO: needs to understand more here
                logger.debug('current rnd_value ' + str(rnd_value))
                logger.debug('current counter ' + str(i))
                logger.debug('self.AutomataList length ' + str(len(self.AutomataList)))

    #---------------------------------------------------------------------------
    # return RGB color string for current cell. no tint value
    # 'XXXXXX'
    #---------------------------------------------------------------------------
    def get_cell_color(self, cell, theme_color):

        color = cell.fill.start_color
        if color.type == "rgb":
            if color.rgb == '00000000':
                return 'FFFFFF'
            else:
                return color.rgb[2:]
        elif color.type == "indexed":
            color_index = color.indexed
            if color_index is None or color_index < len(COLOR_INDEX):
                raise Exception("Invalid indexed color")
            return COLOR_INDEX[color_index][2:]
        elif color.type == "theme":
            return theme_color.theme_and_tint_to_rgb(color.theme, color.tint)
        else:
            raise Exception(f"Other type: {color.type}")

    #---------------------------------------------------------------------------
    # support API. use to extract colors from a excel work book
    #---------------------------------------------------------------------------
    def extract_colors_from_workbook(self, input_file, output_file):
        layout_workbook = openpyxl.load_workbook(input_file)
        theme_color = ThemeColorConverter(layout_workbook)
        # Create a new workbook to save the color data
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Cell Colors"

        # Loop through each sheet in the workbook
        for sheet in layout_workbook.sheetnames:
            ws = layout_workbook[sheet]
        # Loop through each row and column in the sheet
        for row in ws.iter_rows():
            for cell in row:
                newcell = new_ws.cell(row=cell.row, column=cell.column)
                try:
                    color_str = self.get_cell_color(cell, theme_color)
                    newcell.value = color_str
                except:
                    logger.debug('Unexpected color')
                    newcell.value = "UNDEFINIED"

        # Save the new workbook with color information
        new_wb.save(output_file)

    #---------------------------------------------------------------------------
    ##  calcuate pedestrian congestion degree for each emoty cell
    ## aka. all cell in AutomataList
    #---------------------------------------------------------------------------
    def calculatePedCongestion(self):
        # the congestion degree of current empty cell can be calculated by
        #  number of neighbouring cells that are occupied or obstacle / 9
        # have to be called after we have initialized all pedestrian
        for row in range(len(self.LayoutMap)):
            for column in range(len(self.LayoutMap[row])):
                # skip border and obstacle and other cell type, only care about empty spaces
                if self.LayoutMap[row][column].type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                    current_occupied_spaces_around = 0
                    for direction in self.direction_List:
                        current_occupied_spaces_around += self.checkAdjacentCongestions(row, column, direction)
                    self.LayoutMap[row][column].occupied_space_around = current_occupied_spaces_around
                    self.LayoutMap[row][column].ped_congestion = current_occupied_spaces_around / 9.0


    #---------------------------------------------------------------------------
    ##  get the exit ID for current Cell.
    ##  Only works for the cell type = exit_cell defined in the layout input file
    #---------------------------------------------------------------------------
    def getExitIDForCurrentExitCell(self, row, column):
        # check all Von Neumann neighbours of current cell
        # make sure the neighbour is within the range.
        # if the neighbour is empty cell
        # then return the preferred_exit of that empty cell
        exit_id_index = []

        # here we need to check all 8 adjacent cell. no center location
        temp_direction_list = self.direction_List.copy()
        temp_direction_list.remove(self.config.DIRECTION_CENTER)
        for direction in temp_direction_list:
            new_row, new_column = self.getNewPosition(row, column, direction)
            if self.rowOrColumnInRange('row', new_row) and self.rowOrColumnInRange('column', new_column):
                if self.LayoutMap[new_row][new_column].type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                    exit_id_index.append(self.LayoutMap[new_row][new_column].preferred_exit)
                    self.LayoutMap[new_row][new_column].adjacent_to_exit = True

        if exit_id_index is not None:
            if len(exit_id_index) >= 1:
                exit_id = list(set(exit_id_index))[0] #assume we only have 1 values in this case. will not consider exit over 2 zone cases.
            else:
                exit_id = -1
        else:
            exit_id = -1

        # label current exit cell with the correct exit ID.
        ## TODO this might get changed when introduce the exit ID automation algorithm
        self.LayoutMap[row][column].exit_index = exit_id

        return exit_id
    #---------------------------------------------------------------------------
    # Calculate the congestion value for the given position defined by new_row and new_column
    # return 1 if it is border or obstacle or occupied
    #---------------------------------------------------------------------------
    def checkAdjacentCongestions(self, row, column, direction):
        # since the automata cell has relative index, so the neighbour cell's index we have so far are definitely within
        # range. so no further check required.
        ## Current Adacent cell is BORDER or OBSTACLE
        new_row, new_column = self.getNewPosition(row, column, direction)
        if self.LayoutMap[new_row][new_column].type == self.config.LAYOUT_CELL_TYPE_BORDER or \
            self.LayoutMap[new_row][new_column].type == self.config.LAYOUT_CELL_TYPE_OBSTACLE:
            return 1
        elif self.LayoutMap[new_row][new_column].automataIndex != -1: # the adjacent Cell is occupied by an pedestrian
            return 1
        else: # empty space and not occupied by pedestrian
            return 0

    #---------------------------------------------------------------------------
    # based on the paper's algorithm to calculate the static potential matrix
    #---------------------------------------------------------------------------
    def calculateStaticPotentialMatrix(self):
        # create the initial queue of all exit cells
        potential_list = []

        # Here needs to loop entire cell set numer of exit times
        # seems very in effective.
        # TODO Any improvement?
        # for exit_id in range(1, self.config.NUMBER_OF_EXIT + 1):
        for exit_id in range(1, len(self.Exit_Dict) + 1):
            # rest the temp variable & calculation flag
            self.resetCalculationFlag()

            # start the algorithm
            for i in range(len(self.LayoutMap)):
                for j in range(len(self.LayoutMap[i])):
                    # static potential algorithm Step 1. If lattice site (i, j) is occupied by an obstacle,
                    # then set potential of current cell i,j = -1.
                    if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_BORDER or \
                        self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_OBSTACLE:
                        # If this is a border or obstacle, then potentials all sett to -1 with respect of all exit
                        self.LayoutMap[i][j].static_potential = {key : -1 for key in range(1,len(self.Exit_Dict) + 1)}
                        self.LayoutMap[i][j].static_potential_calc_flag = True
                        self.LayoutMap[i][j].temp_potential = -1.0

                    # static potential algorithm Step 2.1. If lattice site (i, j) is occupied by an exit,
                    # then set potential of current cell i,j = 0
                    if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_EXIT:
                        # current exit
                        if self.getExitIDForCurrentExitCell(i,j) == exit_id:
                            # If this is a exit, then potentials all set to 0 with respect of all exit
                            self.LayoutMap[i][j].static_potential = {key : 0 for key in range(1,len(self.Exit_Dict) + 1)}
                            self.LayoutMap[i][j].static_potential_calc_flag = True
                            self.LayoutMap[i][j].temp_potential = 0.0
                            # static potential algorithm Step 2.2. add current exit cell into the queue.
                            potential_list.append(self.LayoutMap[i][j])

            # static potential algorithm Step 2.3 The order of lattice sites increases with their distances.
            potential_list.sort(key=lambda x: x.temp_potential, reverse=False)

            # calculate all cell's potential with respect of current exit id
            while(len(potential_list)>0):
                # static potential algorithm Step 3. For the lattice site (i0, jø) with the lowest distance
                # in the sequence, then check the neighboring lattice site (i, j)
                # in all eight directions and remove the lattice site (io, jo) from the sequence.
                # in this case this will be the first cell in current list as the ascending.
                current_cell = potential_list.pop(0)
                # static potential algorithm Step 3.1 If the feasible distance le i,j of lattice site (i, j)
                # in the horizontal or vertical direction has not been calculated,
                #then let le i,j = le i0,jo + 1.

                # here we need to check all 8 adjacent cell. no center location
                temp_direction_list = self.direction_List.copy()
                temp_direction_list.remove(self.config.DIRECTION_CENTER)
                for direction in temp_direction_list:
                    if direction in [self.config.DIRECTION_DOWN, self.config.DIRECTION_UP, self.config.DIRECTION_LEFT, self.config.DIRECTION_RIGHT]:
                        self.checkAdjacentStaticPotentials(potential_list, current_cell, direction, self.config.HORI_VERT_DISTANCE)
                    else:
                        self.checkAdjacentStaticPotentials(potential_list, current_cell, direction, self.config.DIAGONAL_DISTANCE)

                potential_list.sort(key=lambda x: x.temp_potential, reverse=False)
                #Step 5. If there are no lattice sites for which the feasible distance has not been calculated,
                # then stop; otherwise go to Step 3.

            # Assign the temp potential back to the static potential array
            # the static potential array is a dict and each item there indicates
            # the potentials from current cell with respect to each exit.
            for i in range(len(self.LayoutMap)):
                for j in range(len(self.LayoutMap[i])):
                    # if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                    self.LayoutMap[i][j].static_potential[exit_id] = self.LayoutMap[i][j].temp_potential
            # debug only
            # self.printOutLayout(type='static_potential')

    #---------------------------------------------------------------------------
    # reset the temp_potential and static_potential_calc_flag flag
    #---------------------------------------------------------------------------
    def resetCalculationFlag(self):
        for i in range(len(self.LayoutMap)):
            for j in range(len(self.LayoutMap[i])):
                self.LayoutMap[i][j].static_potential_calc_flag = False
                self.LayoutMap[i][j].temp_potential = 0.0

    #---------------------------------------------------------------------------
    # based on the paper's algorithm to calculate the static potential matrix for an adjacent cell
    #---------------------------------------------------------------------------
    def checkAdjacentStaticPotentials(self, potential_list, current_cell, direction, distance):
        try:
            new_row, new_column = self.getNewPosition(current_cell.row, current_cell.column, direction)
            if self.rowOrColumnInRange('row', new_row) and self.rowOrColumnInRange('column', new_column):
                adjacent_cell = self.LayoutMap[new_row - 1][new_column - 1]
                new_potential = current_cell.temp_potential + distance
                #adjacent cell is not obstacle or boarder and not been calcualted for static potential yet.
                # Step 4. Each lattice site (i, j), whose potential is calculated in Steps 3.1 and 3.2, is added into the sequence
                if not adjacent_cell.static_potential_calc_flag:
                    self.LayoutMap[new_row - 1][new_column - 1].temp_potential = new_potential
                    self.LayoutMap[new_row - 1][new_column - 1].static_potential_calc_flag = True
                    potential_list.append(self.LayoutMap[new_row - 1][new_column - 1])
                    # potential_list.append(adjacent_cell)
                else:
                    if adjacent_cell.static_potential_calc_flag and new_potential < adjacent_cell.temp_potential:
                        self.LayoutMap[new_row - 1][new_column - 1].temp_potential = new_potential
                        potential_list.append(self.LayoutMap[new_row - 1][new_column - 1])
                        # potential_list.append(adjacent_cell)
        except:
            # debug only
            logger.debug('current new_row index  ' + str(new_row - 1))
            logger.debug('current new_column index' + str(new_column - 1))

        # return self.LayoutMap[new_row - 1][new_column - 1]
    #---------------------------------------------------------------------------
    # Check if current Layout is all evacuated.
    #---------------------------------------------------------------------------
    def isEvacuated(self):
        evacuation_flag = False
        if len(self.getPedestrianIndexArray()) == 0:
            evacuation_flag = True
        return evacuation_flag


    #---------------------------------------------------------------------------
    # calculate Avarage Pedestrian Speed
    #---------------------------------------------------------------------------
    def calculateAvaragePedestrianSpeed(self):
        # reset the number of pedestrain in current zone to 0
        # also reset the average speed to 1.2. (Default speed)
        for item in self.evacuation_zone_info:
            item.number_of_pedestrian = 0
            item.average_speed = 0.0

        for cell in self.AutomataList:
            if not cell.is_outside and cell.occupied: # check all pedestrian inside of the field only
                # previous stage spped = (distance / time)
                # distance = how many steps moved in last stage X cell size (assume every step to another cell is same cell size)
                # time = how long it take to move to another cell
                cell.last_speed = (cell.last_step_count * self.config.CELL_SIZE) / \
                (self.config.SIMULATION_CYCLE * (self.config.CELL_SIZE / self.config.NORMAL_PEDESTRIAN_SPEED))

                # will based on current cell's zone id,
                # update the corresponding item in Evacuation Zone list
                # increase the number of pedestrian and add up the average speed.

                self.evacuation_zone_info[cell.current_zone].number_of_pedestrian += 1
                self.evacuation_zone_info[cell.current_zone].average_speed += cell.last_speed

        # Calculate the average speed (total speed/ number of pedestrian)
        for item in self.evacuation_zone_info:
            # if current area has no pedestrain, give the average speed as predefined speed
            if self.evacuation_zone_info[item.assigned_zone_id].number_of_pedestrian == 0:
                self.evacuation_zone_info[item.assigned_zone_id].average_speed = 0.0
            else:
                self.evacuation_zone_info[item.assigned_zone_id].average_speed = \
                    self.evacuation_zone_info[item.assigned_zone_id].average_speed / \
                    self.evacuation_zone_info[item.assigned_zone_id].number_of_pedestrian
        # debug only
        for item in self.evacuation_zone_info:
            logger.debug(item)

    #---------------------------------------------------------------------------
    # calculate the potential of subarea s ? S with respect to exit e ? E
    #---------------------------------------------------------------------------
    def calculatePotentialForSubarea(self):
        subarea_potential_list = []
        # Step 1: for each subarea in the area of pedestrian facility, if e ? E is in
        #  s ? S, F(s,e) = (1 + beta * k(s))(1 + gamma/V(s))
        #  k(s) = number pedestrian in current zone /  zone area
        for exit_id in range(1, len(self.Exit_Dict) + 1):

            # reset the temp potential and calcuation flag
            for item in self.evacuation_zone_info:
                item.temp_potential = 0
                item.potential_calc_flag = False

            for zone_cell in self.evacuation_zone_info:
                if zone_cell.exit_id_for_current_zone == zone_cell.preferred_exit == exit_id:
                    # current zone has an adjacent exit associated and
                    ped_density = zone_cell.number_of_pedestrian / (self.config.EVACUATION_ZONE_DIMENSION * self.config.EVACUATION_ZONE_DIMENSION)
                    try:
                        temp_potential = (1 + self.config.EXIT_CHOICE_MODEL_BETA * ped_density) * \
                                        (1 + self.config.EXIT_CHOICE_MODEL_GAMMA / zone_cell.average_speed)
                    except:
                        logger.info("no pedestrain in current zone. will assign default potential!!")
                        temp_potential = 0.0 # give a default value
                    zone_cell.potential_calc_flag = True
                    zone_cell.temp_potential = temp_potential
                    subarea_potential_list.append(zone_cell)

            # sort the subarea_potential list in increasing order
            subarea_potential_list.sort(key=lambda x: x.temp_potential, reverse=False)

            while(len(subarea_potential_list)>0):
                # Step 2: for subarea s(0) ? S in subarea_potential_list, pop out the smallest in the list (first one)
                # check the neighboring subarea s ? S in all eight directions and remove the subarea s(0) from the
                # subarea_potential list. if the potential of the subarea s has not been dertimined, then calculate
                # the potential of this subarea s based on below 2 cases:
                #       Case 1: if the subarea s is in the giagonal direction of subarea s(0) then
                #       F(s,e) = F(s0,e) + (1+ alpha)(1 + betaK(s))(1 + gamma/V(s))
                #       Case 2: if subarea s is in the horizontal or vertical direction of subarea s(0) then
                #       F(s,e) = F(s0,e) + (1 + betaK(s))(1 + gamma/V(s))
                current_zone_cell = subarea_potential_list.pop(0)
                current_row = current_zone_cell.row
                current_column = current_zone_cell.column

                # here we need to check all 8 adjacent cell. no center location
                temp_direction_list = self.direction_List.copy()
                temp_direction_list.remove(self.config.DIRECTION_CENTER)
                for direction in temp_direction_list:
                    if direction in [self.config.DIRECTION_DOWN, self.config.DIRECTION_UP, self.config.DIRECTION_LEFT, self.config.DIRECTION_RIGHT]:
                        self.checkAdjacentZonePotentials(subarea_potential_list, current_zone_cell, direction, self.config.EXIT_CHOICE_MODEL_ALPHA_NON_DIAGONAL)
                    else:
                        self.checkAdjacentZonePotentials(subarea_potential_list, current_zone_cell, direction, self.config.EXIT_CHOICE_MODEL_ALPHA_DIAGONAL)

                subarea_potential_list.sort(key=lambda x: x.temp_potential, reverse=False)
                #Step 5. If there are no lattice sites for which the feasible distance has not been calculated,
                # then stop; otherwise go to Step 3.

            #assign calculated potential to current exit id
            for item in self.evacuation_zone_info:
                item.potential_dict[exit_id] = item.temp_potential

    #---------------------------------------------------------------------------
    # check a certain direction of current evacuation zone for potentials
    #---------------------------------------------------------------------------
    def checkAdjacentZonePotentials(self, subarea_potential_list, current_cell, direction, alpha):
        current_row = current_cell.row
        current_column = current_cell.column
        new_row, new_column = self.getNewPosition(current_row, current_column, direction)
        try:
            if self.rowOrColumnInRangeForZone('row', new_row, self.evacuation_zone_info) and \
                self.rowOrColumnInRangeForZone('column', new_column, self.evacuation_zone_info):
                # based on the row/column to get the index of entire zone list
                adjacent_zone_cell = self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column]
                ped_density = adjacent_zone_cell.number_of_pedestrian / (self.config.EVACUATION_ZONE_DIMENSION * self.config.EVACUATION_ZONE_DIMENSION)
                try:
                    new_potential = current_cell.temp_potential + \
                                    (1 + alpha) * \
                                    (1 + self.config.EXIT_CHOICE_MODEL_BETA * ped_density ) * \
                                    (1 + self.config.EXIT_CHOICE_MODEL_GAMMA / adjacent_zone_cell.average_speed)
                except:
                    logger.info("no pedestrain in current zone. will assign default potential!!")
                    new_potential = 0.0 # give a default value
                #adjacent cell is not obstacle or boarder and not been calcualted for static potential yet.
                # Step 4. Each lattice site (i, j), whose potential is calculated in Steps 3.1 and 3.2, is added into the sequence
                if not adjacent_zone_cell.potential_calc_flag:
                    self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column].temp_potential = new_potential
                    self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column].potential_calc_flag = True
                    subarea_potential_list.append(self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column])
                else:
                    if adjacent_zone_cell.potential_calc_flag and new_potential < adjacent_zone_cell.temp_potential:
                        self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column].temp_potential = new_potential
                        subarea_potential_list.append(self.evacuation_zone_info[current_cell.number_of_columns * new_row + new_column])
        except:
            # debug only
            logger.debug('current zone new_row index  ' + str(new_row))
            logger.debug('current zone new_column index' + str(new_column))

    #---------------------------------------------------------------------------
    # assign the preferred exit id based on the potentials of current evacuation
    # zone with respect of each exit id.
    # TODO: Maybe this one could be used for initialization of the exit.
    #---------------------------------------------------------------------------
    def assignExitForEvacuationZone(self):
        for zone in self.evacuation_zone_info:
            # current preferred exit will be used as current minimal potential
            # if preferred exit is -1, then will automatically assign the minimal potential
            # associated exit as preferred excit.
            if zone.preferred_exit != -1:
                current_min_potential = zone.potential_dict[zone.preferred_exit]
            else:
                current_min_potential_exit_id = min(zone.potential_dict, key=lambda k: zone.potential_dict[k])
                zone.preferred_exit = current_min_potential_exit_id
                current_min_potential = zone.potential_dict[current_min_potential_exit_id]
            last_potential = current_min_potential
            for exit_id, potential in zone.potential_dict.items():
                potential_diff = last_potential - potential
                if current_min_potential > potential and potential_diff > self.config.EXIT_CHOICE_MODEL_THETA:
                    current_min_potential = potential
                    zone.preferred_exit = exit_id
        #debug only
        pass

        # Here we also needs to assign the preferred_exit in the LayoutMap
        for i in range(len(self.LayoutMap)):
            for j in range(len(self.LayoutMap[i])):
                # we have an assigned evacuation zone based on initial potential already
                # here only assign the preferred_exit for those cell that occupied
                if self.LayoutMap[i][j].assigned_evacuation_zone != -1:
                    self.LayoutMap[i][j].preferred_exit = self.evacuation_zone_info[self.LayoutMap[i][j].assigned_evacuation_zone].preferred_exit

    #---------------------------------------------------------------------------
    # helper function. by given check type (row or column), check the value is in range
    #---------------------------------------------------------------------------
    def rowOrColumnInRange(self, type, value):
        if type.lower() == 'row':
            return value > self.min_row - 1 and value < self.max_row
        if type.lower() == 'column':
            return value > self.min_column - 1 and value < self.max_column
        return False # fallback to false to be safe

    #---------------------------------------------------------------------------
    # helper function. by given check type (row or column), check the value is in range based on the given source
    #---------------------------------------------------------------------------
    def rowOrColumnInRangeForZone(self, type, value, source):
        # Here the source is the array of a certain item.
        # we just store the 2 dimension array into 1 dimension list
        # so assume the last item of this list has the max_row/max_column and
        # the first item of this list has the min_row/min_columen
        if type.lower() == 'row':
            return value >= source[0].row and value <= source[-1].row
        if type.lower() == 'column':
            return value >= source[0].column and value <= source[-1].column
        return False # fallback to false to be safe

    #---------------------------------------------------------------------------
    # helper function. print out layout for debug only
    #---------------------------------------------------------------------------
    def printOutLayout(self, type='value'):
        ## print out the layout array per value
        in_range = False
        for i in range(len(self.LayoutMap)):
            for j in range(len(self.LayoutMap[i])):
                if self.LayoutMap[i][j].row > self.min_row - 1 and  \
                    self.LayoutMap[i][j].column > self.min_column - 1 and \
                    self.LayoutMap[i][j].row < self.max_row and \
                    self.LayoutMap[i][j].column < self.max_column:
                    # print(" Current Row : " + str(self.LayoutMap[i][j].row) +
                    #     " Current Column : " + str(self.LayoutMap[i][j].column))
                    if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                        if type == 'assigned_evacuation_zone':
                            logger.info('|' + str(self.LayoutMap[i][j].assigned_evacuation_zone), end='')
                        elif type == 'static_potential':
                            for exit in range(1, len(self.Exit_Dict) + 1):
                                logger.info('| (' + str(exit) + ')' + str(self.LayoutMap[i][j].static_potential[exit]), end='')
                        elif type == 'preferred_exit':
                            logger.info('|' + str(self.LayoutMap[i][j].preferred_exit), end='')
                        elif type == 'ped_congestion':
                            logger.info('|' + str(self.LayoutMap[i][j].ped_congestion), end='')
                        else: # default to value
                            logger.info('|' + str(self.LayoutMap[i][j].value), end='')
                    else:
                        logger.info('|' + 'B', end='')
                    in_range = True
            if in_range:
                logger.info('|')
                logger.info('---------------------------------------------------------------------')

    #---------------------------------------------------------------------------
    # Get current index array from Automata List that is occupied by pedestrian
    #---------------------------------------------------------------------------
    def getPedestrianIndexArray(self):
        idx_array = []
        for i in range(len(self.AutomataList)):
            if self.AutomataList[i].occupied and not self.AutomataList[i].is_outside:
                idx_array.append(i)

        return idx_array


    #---------------------------------------------------------------------------
    # Calcuate the Potential of give cell with respect of the new row/column
    # corresponding to a specific exit
    #---------------------------------------------------------------------------
    def calculateDynamicPotential(self, current_LayoutCell, direction, exit_id):
        # The potential of each lattice site is used to reflect the total effects of the dynamic
        # potential and the static potential of each lattice site. Therefore, the potential f(e(i,j))
        # can be expressed as follows:

        # f(e(i,j)) = do(i,j)/n(i,j) + ?(l(e(i,j)) - l(e(i0,j0))
        real_index_row = current_LayoutCell.row - 1
        real_index_column = current_LayoutCell.column - 1

        new_row, new_column = self.getNewPosition(real_index_row, real_index_column, direction)
        if self.rowOrColumnInRange('row', new_row) and self.rowOrColumnInRange('column', new_column):
            new_LayoutCell = self.LayoutMap[new_row][new_column]
            return self.config.PEDESTRIAN_MOVEMENT_MODEL_DELTA * current_LayoutCell.ped_congestion + \
                    self.config.PEDESTRIAN_MOVEMENT_MODEL_PHI * (new_LayoutCell.static_potential[exit_id] - current_LayoutCell.static_potential[exit_id])
        else:
            # out of range cell
            # return some dummy value here
            logger.info('Out of Range. Use default dummy potential')
            return -100.0

    #---------------------------------------------------------------------------
    # get the adjust row/column based on the direction and return the properly
    # index  can idenfity the layout cell from Layout Map.
    # will consider the border/obstacle
    #---------------------------------------------------------------------------
    def getAdjustedLayoutIndex(self, row, column, direction):
        pass



    #---------------------------------------------------------------------------
    # Calcuate the Transition Probability for current cell and store them into a
    # dict of all Transition Probabilities.
    #---------------------------------------------------------------------------
    def calculateTransitionProbability(self, shuffled_idx_array):
        for idx in shuffled_idx_array:

            DynamicTransitionProbability_Dict = {}
            DynamicPotential_Dict = {}
            DynamicCellAccessibility_Dict = {}

            current_CA = self.AutomataList[idx]
            current_idx_row = current_CA.x - 1        # real index that can idenfity from Layout Map 2 dimension array
            current_idx_column = current_CA.y - 1     # real index that can idenfity from Layout Map 2 dimension array
            current_LayoutCell = self.LayoutMap[current_idx_row][current_idx_column]
            ## CAUTION: the row/column here is the index one. not same as the row/column information inside of the Layoutcell object
            current_exit_id = current_LayoutCell.preferred_exit
            # # Assign the current_exit in AutomataList as well. Might not needed?
            # current_CA.current_exit = current_LayoutCell.preferred_exit
            # if current_exit_id == -1:
            #     pass

            # For debug only
            logger.info(f"current Pedestrian cell [{idx}]: {self.AutomataList[idx]} ")

            if current_LayoutCell.adjacent_to_exit:
                logger.info(f'Pedestrian {idx} exit the facility through exit : {current_exit_id}')
                # here the current Cell should get the closest exit's row and column info.
                exit_row, exit_column = self.getClosestExitCoordination(current_LayoutCell)
                current_CA.move_history.append(tuple([exit_row,exit_column]))
                current_LayoutCell.automataIndex = -1
                current_CA.is_outside = True
                current_CA.occupied  = False
            else:
                # Calculate the Transition probability of each direction around current cell.
                # get dynamic potentials adn accessibility of each cell around current cell
                for direction in self.direction_List:
                    DynamicPotential_Dict[direction] = self.calculateDynamicPotential(current_LayoutCell, direction, current_exit_id)
                    DynamicCellAccessibility_Dict[direction] = self.checkAdjacentCellStatus(current_LayoutCell, direction)

                # sort the dict so we can perform python list comprehension below
                DynamicCellAccessibility_Dict = dict(sorted(DynamicCellAccessibility_Dict.items()))
                DynamicPotential_Dict = dict(sorted(DynamicPotential_Dict.items()))

                ## Main transition Probability formula
                for key, value in DynamicCellAccessibility_Dict.items():
                    DynamicTransitionProbability_Dict[key] = value / \
                        sum(math.exp(-self.config.PEDESTRIAN_MOVEMENT_MODEL_EPSILON *
                                     (potential - DynamicPotential_Dict[key])) for potential in DynamicPotential_Dict.values())

                # update current Cell to move to the specific direction based on the Transtion Probability dict.
                # the larger probability direction will be chose.
                ## TODO: Still need to understand the original C++ code's algorithm
                max_key = max(DynamicTransitionProbability_Dict.items(), key = operator.itemgetter(1))[0]
                max_value = max(DynamicTransitionProbability_Dict.values())
                new_row, new_column = self.getNewPosition(current_idx_row, current_idx_column, max_key)


                # add the move history into automata celluar
                current_CA.move_history.append(tuple([new_row + 1,new_column + 1]))
                # reassign the new x, y in current CA.
                current_CA.x = new_row + 1
                current_CA.y = new_column + 1

                # to set the new layout cell's automataIndex to current Automata's Idx.
                # to set the original layout cell's automataIndex to -1.
                # incrase the last_step_count if current pedestrian moved to another location.
                if new_row != current_idx_row or new_column != current_idx_column: # not stay in same position
                    self.LayoutMap[new_row][new_column].automataIndex = current_LayoutCell.automataIndex
                    current_LayoutCell.automataIndex = -1
                    # TODO: if we assume we can move both veritical/horizontal & diagonal, then the step count should also
                    # consider that. if we move diagonal then last step count should be sqrt(2)?
                    if max_key in [self.config.DIRECTION_DOWN, self.config.DIRECTION_LEFT, self.config.DIRECTION_UP, self.config.DIRECTION_RIGHT]:
                        current_CA.last_step_count += 1
                    else:
                        current_CA.last_step_count += math.sqrt(2)

                # Log info: will use the real row/column value but not the index here.
                logger.debug(f'Pedestrian {idx} move : {max_key} from [row : {current_idx_row + 1}, column : {current_idx_column + 1}] to [row : {new_row + 1}, column : {new_column + 1}] with probability {max_value:.5%}')

    #---------------------------------------------------------------------------
    # Get new row, new column index based on current cell. will find out the most
    # closest exit and return that exit cell's row & column
    #---------------------------------------------------------------------------
    def getClosestExitCoordination(self, cell):
        distance_dict={}
        temp_direction_list = self.direction_List.copy()
        temp_direction_list.remove(self.config.DIRECTION_CENTER)
        for direction in temp_direction_list:
            new_row, new_column = self.getNewPosition(cell.row, cell.column, direction)
            # if self.rowOrColumnInRange('row', new_row) and self.rowOrColumnInRange('column', new_column):
            if self.LayoutMap[new_row - 1][new_column - 1].type == self.config.LAYOUT_CELL_TYPE_EXIT:
                if direction in [self.config.DIRECTION_DOWN, self.config.DIRECTION_LEFT, self.config.DIRECTION_UP, self.config.DIRECTION_RIGHT]:
                    distance_dict[tuple([new_row, new_column])] = 1 # vertical or horizontal
                else:
                    distance_dict[tuple([new_row, new_column])] = math.sqrt(2) # diagonal
        # if len(distance_dict) == 0:
        #     pass
        exit_row, exit_column = min(distance_dict, key=distance_dict.get)
        # will return the actual row/column info of the new cell. not there idx
        return exit_row, exit_column

    #---------------------------------------------------------------------------
    # Get new row, new column index based on current row, column information plus the direction
    #---------------------------------------------------------------------------
    def getNewPosition(self, current_row, current_column, direction):
        new_row = new_column = -1
        if direction == self.config.DIRECTION_UP:
            new_row = current_row - 1
            new_column = current_column
        if direction == self.config.DIRECTION_DOWN:
            new_row = current_row + 1
            new_column = current_column
        if direction == self.config.DIRECTION_LEFT:
            new_row = current_row
            new_column = current_column - 1
        if direction == self.config.DIRECTION_RIGHT:
            new_row = current_row
            new_column = current_column + 1
        if direction == self.config.DIRECTION_DOWNLEFT:
            new_row = current_row + 1
            new_column = current_column - 1
        if direction == self.config.DIRECTION_DOWNRIGHT:
            new_row = current_row + 1
            new_column = current_column + 1
        if direction == self.config.DIRECTION_UPLEFT:
            new_row = current_row - 1
            new_column = current_column - 1
        if direction == self.config.DIRECTION_UPRIGHT:
            new_row = current_row - 1
            new_column = current_column + 1
        if direction == self.config.DIRECTION_CENTER:
            new_row = current_row
            new_column = current_column

        return new_row, new_column




    #---------------------------------------------------------------------------
    # check whether the adjacent cell of a given cell is empty or not.
    # Only return 1 if the adjacent cell is not occupied
    # return 0 if the adjacent cell is either occupied or obstacle or border or other
    # not moveable cell type.
    #---------------------------------------------------------------------------
    def checkAdjacentCellStatus(self, cell, direction):
        status = 0
        real_index_row = cell.row - 1
        real_index_column = cell.column - 1
        new_row, new_column = self.getNewPosition(real_index_row, real_index_column, direction)
        if self.rowOrColumnInRange('row', new_row) and self.rowOrColumnInRange('column', new_column):
            if self.LayoutMap[new_row][new_column].type in \
                        [self.config.LAYOUT_CELL_TYPE_BORDER,
                            self.config.LAYOUT_CELL_TYPE_OBSTACLE] or \
                self.LayoutMap[new_row][new_column].automataIndex != -1:
                status = 0
            else:
                status = 1

        # CENTER always give 1
        if direction == self.config.DIRECTION_CENTER:
            status = 1

        return status

    #---------------------------------------------------------------------------
    # with passed in buffered DC, based on current LayoutMap, repaint the DC.
    #---------------------------------------------------------------------------
    def refresh_layoutmap(self, canvas):
        logger.info("Refreshing Screen...")
        canvas.clear() # clear current layout

        for row in self.LayoutMap[self.min_row - 1:]:
            for cell in row[self.min_column - 1:]:
                # for border/obstacle/exit, draw solid rectangle
                if cell.type == self.config.LAYOUT_CELL_TYPE_BORDER or \
                    cell.type == self.config.LAYOUT_CELL_TYPE_OBSTACLE or \
                    cell.type == self.config.LAYOUT_CELL_TYPE_EXIT:
                    pen = QPen(QColor("BLACK"), 1, Qt.SolidLine)  # Blue color, 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(cell.color)), Qt.SolidPattern)  # Red solid fill
                    rect = QGraphicsRectItem(QRectF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    self.config.LAYOUT_CELL_SIZE,
                                    self.config.LAYOUT_CELL_SIZE))  # Rectangle with (x, y, width, height)
                    rect.setPen(pen)
                    rect.setBrush(brush)
                    canvas.addItem(rect)
                else:
                    verticle_line = QGraphicsLineItem(QLineF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    cell.column * self.config.LAYOUT_CELL_SIZE + self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE))
                    horizontal_line = QGraphicsLineItem(QLineF(cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE,
                                    cell.column * self.config.LAYOUT_CELL_SIZE,
                                    cell.row * self.config.LAYOUT_CELL_SIZE + self.config.LAYOUT_CELL_SIZE))
                    # Create a pen and brush
                    pen = QPen(QColor("grey"), 1)  # Black pen with 1px width, solid line
                    verticle_line.setPen(pen)
                    horizontal_line.setPen(pen)
                    canvas.addItem(verticle_line)
                    canvas.addItem(horizontal_line)

        for i in range(len(self.AutomataList)):
            # here adjust the start size and cell size to make the graph more precise
            # We draw the pedestrian inside of each cell. Not cover the border.
            if self.AutomataList[i].occupied: # only setup the spot that is occupied. (not border/exit/obstacle)
                ellipse = QGraphicsEllipseItem(QRectF((self.AutomataList[i].y) * self.config.LAYOUT_CELL_SIZE + 1,
                                    (self.AutomataList[i].x) * self.config.LAYOUT_CELL_SIZE + 1,
                                    self.config.LAYOUT_CELL_SIZE - 2,
                                    self.config.LAYOUT_CELL_SIZE - 2))
                if self.AutomataList[i].velocity_mode == 2: # senior people
                    pen = QPen(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_SENIOR_COLOR)), 1, Qt.SolidLine)  # Black pen with 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_SENIOR_COLOR)), Qt.SolidPattern)  # Red solid brush
                else:
                    pen = QPen(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_NON_SENIOR_COLOR)), 1, Qt.SolidLine)  # Black pen with 3px width, solid line
                    brush = QBrush(QColor(*ThemeColorConverter.hex_to_rgb(self.config.LAYOUT_NON_SENIOR_COLOR)), Qt.SolidPattern)  # Red solid brush
                ellipse.setPen(pen)
                ellipse.setBrush(brush)
                canvas.addItem(ellipse)

    #---------------------------------------------------------------------------
    # Find out all exits of current layout
    # should be call after we have the layout loaded from excel file.
    # 1. looking for exit_cell.
    # 2. once find out the one exit_cell, check all directions for adjacent cell
    # 3. if adjacent cell is also exit_cell, add it into one exit cell list untill
    #    all cells in current exit cell list are all checked that no other adjacent cells
    #    will be added into current cell.
    # 4. continute this process to check other cells in current layout map. if other cells are
    #    already in exiting exit cell list group, then ignore.
    # 5. stop when all cells in current layout map are visitied.
    #---------------------------------------------------------------------------
    def initializeExitAndEvacuationZone(self):
        # find out all exits from current layout map
        temp_exit_list=[]
        for i in range(len(self.LayoutMap)):
            for j in range(len(self.LayoutMap[i])):
                if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_EXIT and not self.LayoutMap[i][j].assigned_to_exit_flag:
                    current_exit_list = self.findAllAdjacentExitCell(self.LayoutMap[i][j].row, self.LayoutMap[i][j].column)
                    temp_exit_list.append(current_exit_list)
        # all exit has been assigned with id start from 1.
        exit_id = 1
        for item in temp_exit_list:
            self.Exit_Dict[exit_id] = item
            for cell in item:
                cell.exit_index = exit_id
            exit_id += 1

        ## we have all cells visited. check the Exit_Dict and assign initial exit for
        ## each cell again and also assign exit for evacuation zone.
        ## TODO: in a different layout. how to assign the exit for each zone?
        ##      1. based on the layout dimension derived the size of evacuation zone
        ##      2. calculate how many exit in current layout (based on connected exit cell?)
        ##      3. check every layout cell if it is adjacent to an exit.
        ##      4. assign adjacent_to_exit attribute of each layout cell
        ##      5. assign current_exit attribute for evacuation zone cell if the layout cell in current zone has adjacent_to_exit = True
        ##      6. use LFS to calculate all evacuation zone's preferred_exit
        ## Here we need to check all include the border...
        ## TODO need to check all for loop to make sure it is correct.
        for i in range(len(self.LayoutMap)):
            for j in range(len(self.LayoutMap[i])):
                if self.LayoutMap[i][j].type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                    current_row = self.LayoutMap[i][j].row
                    current_column = self.LayoutMap[i][j].column
                    # here we need to check all 8 adjacent cell. no center location
                    temp_direction_list = self.direction_List.copy()
                    temp_direction_list.remove(self.config.DIRECTION_CENTER)
                    for direction in temp_direction_list:
                        # Don't need to check row/column within the range at this moment since we are checking Exit here.
                        # use try/except to make sure the program can work
                        new_row, new_column = self.getNewPosition(current_row, current_column, direction)
                        try:
                            if self.LayoutMap[new_row - 1][new_column - 1].type == self.config.LAYOUT_CELL_TYPE_EXIT:
                                # we find out an adjacent cell of current cell that is EXIT.
                                # mark currnet cell adjacent to exit = True
                                self.LayoutMap[i][j].adjacent_to_exit = True
                                self.LayoutMap[i][j].preferred_exit = self.LayoutMap[new_row - 1][new_column - 1].exit_index

                                # for the evacuation zone that current cell belongs to
                                # assign the same preferred exit to the entire zone.
                                for zone in self.evacuation_zone_info:
                                    if zone.assigned_zone_id == self.LayoutMap[i][j].assigned_evacuation_zone:
                                        zone.preferred_exit = self.LayoutMap[new_row - 1][new_column - 1].exit_index
                                break
                        except:
                            logger.debug("row/column out of range. ignored")

        # we have all evacuation zone that have direct adjacent exit assigned with that exit already.
        # needs to assign other evacuation zone
        self.calculateStaticDistanceMatrixForEvacuationZone()

        for zone in self.evacuation_zone_info:
            # Find the smallest distance
            min_value = min(zone.static_distance_dict.values())

            # Collect all keys with the smallest value
            min_keys = [k for k, v in zone.static_distance_dict.items() if v == min_value]

            # Randomly pick one key from the keys with the smallest value
            random_key = random.choice(min_keys)
            zone.preferred_exit = random_key

        # once we have all evacuation zone been assigned, will assign all cell that belongs to that evacuation zone with same
        # exit id.
        for row in self.LayoutMap[self.min_row - 1:]:
            for cell in row[self.min_column - 1:]:
                if cell.type == self.config.LAYOUT_CELL_TYPE_EMPTY_SPACE:
                    cell.preferred_exit = self.evacuation_zone_info[cell.assigned_evacuation_zone].preferred_exit

    #---------------------------------------------------------------------------
    # calculate static distance for each zone aganist each exit
    #---------------------------------------------------------------------------
    def calculateStaticDistanceMatrixForEvacuationZone(self):
        zone_list = []
        # loop through all exits in current layout.
        # Step 1: if current zone(i,j) is adjacent to current exit then l(i,j) = 0
        # Step 2: add current zone into an ordered sequences that needs to be checked.
        #         the order of the zone list increases with their distance towards current exit
        # Step 3: for the zone cell with lowest distance in the sequence, check all neighboring zone
        #         and remove current zone from the sequences.
        # Step 4: if the adjacent zone in the horizontal or vertical direction has not been calculated
        #         then l(i,j) = l(i0,j0) + 1. if the adjancet zong is in diagonal direction, then
        #          l(i,j) = l(i0, j0) + aqrt(2)
        # Step 5: add the zone whose static distance has been calculated into the sequences.
        # Step 6: if no zone that the static distance has not been calculated, then stop, otherwise
        #         continue to step 4.
        for exit_id in self.Exit_Dict.keys():

            # reset the temp distance and distance calcuation flag
            for item in self.evacuation_zone_info:
                item.temp_distance = 0
                item.distance_calc_flag = False

            for zone_cell in self.evacuation_zone_info:
                if zone_cell.obstacle_zone_flag:
                    temp_distance = -1
                    zone_cell.distance_calc_flag = True
                if zone_cell.preferred_exit == exit_id:
                    # current zone has an adjacent exit associated
                    temp_distance = 0
                    zone_cell.distance_calc_flag = True
                    zone_cell.temp_distance = temp_distance
                    zone_list.append(zone_cell)

            # sort the subarea_potential list in increasing order
            zone_list.sort(key=lambda x: x.temp_distance, reverse=False)

            while(len(zone_list)>0):
                current_zone_cell = zone_list.pop(0)
                # here we need to check all 8 adjacent zone. no center location
                temp_direction_list = self.direction_List.copy()
                temp_direction_list.remove(self.config.DIRECTION_CENTER)
                for direction in temp_direction_list:
                    if direction in [self.config.DIRECTION_DOWN, self.config.DIRECTION_UP, self.config.DIRECTION_LEFT, self.config.DIRECTION_RIGHT]:
                        self.checkAdjacentZoneDistance(zone_list, current_zone_cell, direction, self.config.HORI_VERT_DISTANCE)
                    else:
                        self.checkAdjacentZoneDistance(zone_list, current_zone_cell, direction, self.config.DIAGONAL_DISTANCE)

                zone_list.sort(key=lambda x: x.temp_distance, reverse=False)
                #Step 5. If there are no lattice sites for which the feasible distance has not been calculated,
                # then stop; otherwise go to Step 3.

            #assign calculated potential to current exit id
            for item in self.evacuation_zone_info:
                item.static_distance_dict[exit_id] = item.temp_distance

    #---------------------------------------------------------------------------
    # check a certain direction of current evacuation zone for potentials
    #---------------------------------------------------------------------------
    def checkAdjacentZoneDistance(self, zone_list, current_zone_cell, direction, distance):
        current_row = current_zone_cell.row
        current_column = current_zone_cell.column
        new_row, new_column = self.getNewPosition(current_row, current_column, direction)
        try:
            if self.rowOrColumnInRangeForZone('row', new_row, self.evacuation_zone_info) and \
                self.rowOrColumnInRangeForZone('column', new_column, self.evacuation_zone_info):
                # based on the row/column to get the index of entire zone list
                adjacent_zone_cell = self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column]
                new_distance = current_zone_cell.temp_distance + distance
                if not adjacent_zone_cell.distance_calc_flag:
                    self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column].temp_distance = new_distance
                    self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column].distance_calc_flag = True
                    zone_list.append(self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column])
                else:
                    if adjacent_zone_cell.distance_calc_flag and new_distance < adjacent_zone_cell.temp_distance:
                        self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column].temp_distance = new_distance
                        zone_list.append(self.evacuation_zone_info[current_zone_cell.number_of_columns * new_row + new_column])
        except:
            # debug only
            logger.debug('calculating zone evacuation exit : current zone new_row index  ' + str(new_row))
            logger.debug('calculating zone evacuation exit' + str(new_column))

    #---------------------------------------------------------------------------
    # Find out all exits cells based on current current_exitcell_list
    #---------------------------------------------------------------------------
    def findAllAdjacentExitCell(self, current_row, current_column):
        # here we cannot use the rowOrColumn in range method to check. since all exit are consider
        # boundary cell already.
        current_exitcell_list = [self.LayoutMap[current_row - 1][current_column - 1]]
        self.LayoutMap[current_row - 1][current_column - 1].assigned_to_exit_flag = True # mark current exit visited.
        result_exitcell_list = []
        while len(current_exitcell_list) > 0:
            current_cell = current_exitcell_list.pop(0)
            result_exitcell_list.append(current_cell)
            for direction in self.direction_List:
                new_row, new_column = self.getNewPosition(current_cell.row, current_cell.column, direction)
                try:
                    if self.LayoutMap[new_row - 1][new_column - 1] is not None: # if we can access the cell in current direction
                        if self.LayoutMap[new_row - 1][new_column - 1].type == self.config.LAYOUT_CELL_TYPE_EXIT and \
                            not self.LayoutMap[new_row - 1][new_column - 1].assigned_to_exit_flag:
                            current_exitcell_list.append(self.LayoutMap[new_row - 1][new_column - 1])
                        self.LayoutMap[new_row - 1][new_column - 1].assigned_to_exit_flag = True
                    else:
                        pass
                        logger.debug("Don't check off boundary now. just check if we can have adjacent exit cell")
                except:
                    logger.debug("Don't check off boundary now. just check if we can have adjacent exit cell")
        return result_exitcell_list
